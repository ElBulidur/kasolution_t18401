import os
from abc import ABC, abstractclassmethod

arquivo = "logs.log"
pasta = "logs"


# print( os.path.exists(arquivo) ) # Verifica se o arquivo existe.
# print( os.path.isdir(pasta) ) # Verifica se a pasta existe.

# print( os.getcwd() ) # mostra aonde esta sendo executado
# os.chdir("logs") # navegar para o caminho especifico
# os.mkdir("dados") #  Criar pasta
# os.rmdir("dados") # Deletar

# os.system("touch logs.log") # usando comando linux para criar arquivos
# os.system("rm logs.log") # usando comando linux para deletar arquivos
# os.system("clear") # limpar terminal

# print( os.listdir() )

conteudo = os.listdir()

arquivos_python = []

# for x in conteudo:
#     if os.path.isdir(x):
#         print(f"{x} é uma pasta")
#     else:
#         print(f"{x} é um arquivo com extenção {x.split(".")[1]}")
#         if x.split(".")[1] =="py":
#             arquivos_python.append(x)
            
            
# print(arquivos_python)

nome = "estoque.txt"

# arquivo = open("estoque.txt", "w", encoding="utf-8")

# arquivo.write("Produto, Preço, Qtd\n")
# arquivo.write("Mouse, 42.54, 10\n")

# arquivo.close()

# with open("estoque.txt", "w", encoding="utf-8") as arquivo:
#     arquivo.write("Produto, Preço, Qtd\n")
#     arquivo.write("LAPIS, 10.54, 10\n")
#     arquivo.write("Mouse, 42.54, 8\n")
#     arquivo.write("BATATA, 2.54, 2\n")
#     arquivo.write("CEBOLA, 5.54, 13\n")
#     print("Arquivo criado com sucesso!!!")

# from dotenv import load_dotenv

# load_dotenv()

# print(os.getenv("DB_NAME"))


#LEITURA DE ARQUIVO
# with open("estoque.txt", "r", encoding="utf-8") as arquivo:
#     for linha in arquivo:
#         produto, preco, qtd = linha.split(", ")
#         print(f"{produto:<10} | Preço: R$ {preco} | qtd: {qtd}")

        
# with open("estoque.txt", "a", encoding="utf-8") as arquivo:
#     arquivo.write("MAÇA, 10.54, 10\n")


# CSV
import csv


dados = [
    ["Produto", "Preço", "QTD"],
    ["Mouse", 48.15, 5],
    ["Teclado", 12.15, 15],
    ["Laranja", 4.15, 50]
]

# arquivo = open("estoque.csv", "w", newline="", encoding="utf-8")
# escrever = csv.writer(arquivo)

# escrever.writerows(dados)
# arquivo.close()


arquivo = open("data_enem.csv", "r", newline="")

leitor = csv.reader(arquivo)

# for linha in leitor:
#     print(linha)
#     break
    
arquivo.close()


# EXCEL
from openpyxl import Workbook, load_workbook
from datetime import datetime

# CRIAR EXCEL
wb = Workbook()
ws = wb.active

ws.title = "Estoque"

ws.append(["Produto", "Preço"])

ws.append(["Mouse", 45.42])

wb.save("estoque.xlsx")



# LER EXCEL
wb = load_workbook("data_enem.xlsx")
ws = wb["PARTICIPANTES_2024"]

for linha in ws.iter_rows(values_only=True):
    print(linha)
    break

wb.close()





class Relatorio(ABC):
    pass


class RelatorioEstoque(Relatorio):
    pass

    def criar_em_txt():
        pass
    
    def criar_em_excel(self, cabecalho:list, dados: list):
        wb = Workbook()
        ws = wb.active

        ws.title = "Estoque"

        ws.append(cabecalho)
        
        for linha in dados:
            ws.append(linha)

        wb.save(f"relatorio_estoque_{datetime.now().strftime("%d.%m.%Y")}.xlsx")
        


dados = [
    ["Produto", "Preço", "QTD"],
    ["Mouse", 48.15, 5],
    ["Teclado", 12.15, 15],
    ["Laranja", 4.15, 50]
]

      
rel = RelatorioEstoque()

rel.criar_em_excel(dados[0], dados[1:3])

